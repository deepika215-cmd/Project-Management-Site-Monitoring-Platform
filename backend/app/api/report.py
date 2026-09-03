from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from app.database.database import get_db

from app.models.report import Report
from app.models.project import Project
from app.models.daily_progress import DailyProgress
from app.models.weekly_progress import WeeklyProgress
from app.models.delay_record import DelayRecord
from app.models.project_milestone import ProjectMilestone

from app.schemas.report_schema import (
    ReportCreate,
    ReportResponse,
)

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import os


router = APIRouter(
    prefix="/report",
    tags=["Report"],
)


# ============================================================
# CREATE REPORT
# ============================================================

@router.post("/", response_model=ReportResponse)
def create_report(
    report: ReportCreate,
    db: Session = Depends(get_db),
):
    new_report = Report(
        **report.model_dump()
    )

    db.add(new_report)
    db.commit()
    db.refresh(new_report)

    return new_report


# ============================================================
# GET ALL REPORTS
# ============================================================

@router.get("/", response_model=list[ReportResponse])
def get_reports(
    db: Session = Depends(get_db),
):
    return db.query(Report).all()


# ============================================================
# GET REPORT BY ID
# ============================================================

@router.get("/{report_id}", response_model=ReportResponse)
def get_report(
    report_id: int,
    db: Session = Depends(get_db),
):
    report = (
        db.query(Report)
        .filter(Report.id == report_id)
        .first()
    )

    if not report:
        raise HTTPException(
            status_code=404,
            detail="Report not found",
        )

    return report


# ============================================================
# UPDATE REPORT
# ============================================================

@router.put("/{report_id}", response_model=ReportResponse)
def update_report(
    report_id: int,
    updated_report: ReportCreate,
    db: Session = Depends(get_db),
):
    report = (
        db.query(Report)
        .filter(Report.id == report_id)
        .first()
    )

    if not report:
        raise HTTPException(
            status_code=404,
            detail="Report not found",
        )

    for key, value in updated_report.model_dump().items():
        setattr(report, key, value)

    db.commit()
    db.refresh(report)

    return report


# ============================================================
# DELETE REPORT
# ============================================================

@router.delete("/{report_id}")
def delete_report(
    report_id: int,
    db: Session = Depends(get_db),
):
    report = (
        db.query(Report)
        .filter(Report.id == report_id)
        .first()
    )

    if not report:
        raise HTTPException(
            status_code=404,
            detail="Report not found",
        )

    db.delete(report)
    db.commit()

    return {
        "message": "Report deleted successfully"
    }


# ============================================================
# GENERATE PROJECT PDF REPORT
# ============================================================

@router.get("/project/{project_id}/pdf")
def generate_project_pdf(
    project_id: int,
    db: Session = Depends(get_db),
):
    # --------------------------------------------------------
    # Get project
    # --------------------------------------------------------

    project = (
        db.query(Project)
        .filter(Project.id == project_id)
        .first()
    )

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found",
        )

    # --------------------------------------------------------
    # Get Module 3 data
    # --------------------------------------------------------

    daily_progress = (
        db.query(DailyProgress)
        .filter(DailyProgress.project_id == project_id)
        .order_by(DailyProgress.report_date)
        .all()
    )

    weekly_progress = (
        db.query(WeeklyProgress)
        .filter(WeeklyProgress.project_id == project_id)
        .order_by(WeeklyProgress.week_start)
        .all()
    )

    delays = (
        db.query(DelayRecord)
        .filter(DelayRecord.project_id == project_id)
        .order_by(DelayRecord.delay_date)
        .all()
    )

    milestones = (
        db.query(ProjectMilestone)
        .filter(ProjectMilestone.project_id == project_id)
        .order_by(ProjectMilestone.due_date)
        .all()
    )

    # --------------------------------------------------------
    # Create output directory
    # --------------------------------------------------------

    output_directory = "generated_reports"

    os.makedirs(
        output_directory,
        exist_ok=True
    )

    file_name = (
        f"project_{project_id}_progress_report.pdf"
    )

    file_path = os.path.join(
        output_directory,
        file_name
    )

    # --------------------------------------------------------
    # Create PDF
    # --------------------------------------------------------

    document = SimpleDocTemplate(
        file_path,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    story = []

    # --------------------------------------------------------
    # Title
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "BuildTrack - Project Progress Report",
            styles["Title"],
        )
    )

    story.append(
        Spacer(1, 15)
    )

    # --------------------------------------------------------
    # Project Information
    # --------------------------------------------------------

    project_data = [
        ["Project Name", project.project_name or ""],
        ["Project Code", project.project_code or ""],
        ["Category", project.project_category or ""],
        ["Location", project.location or ""],
        ["Start Date", str(project.start_date or "")],
        ["End Date", str(project.end_date or "")],
        ["Budget", str(project.budget or "")],
        ["Status", project.status or ""],
    ]

    project_table = Table(
        project_data,
        colWidths=[130, 350],
    )

    project_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ])
    )

    story.append(project_table)

    story.append(
        Spacer(1, 20)
    )

    # ========================================================
    # DAILY PROGRESS
    # ========================================================

    story.append(
        Paragraph(
            "Daily Progress Reports",
            styles["Heading2"],
        )
    )

    if daily_progress:

        daily_data = [
            [
                "Date",
                "Category",
                "Activity",
                "Completion %",
                "Contractor",
                "Workers",
                "Machinery",
                "Materials",
                "Weather",
                "Delay",
            ]
        ]

        for item in daily_progress:

            daily_data.append([
                str(item.report_date or ""),
                item.work_category or "",
                item.activity or "",
                str(item.completion_percentage or 0),
                item.contractor_name or "",
                f"P:{item.workers_present or 0} / A:{item.workers_absent or 0}",
                item.machinery_used or "",
                item.materials_used or "",
                item.weather or "",
                f"{item.delay_hours or 0} hrs",
            ])

        daily_table = Table(
            daily_data,
            repeatRows=1,
            colWidths=[
                65,
                75,
                100,
                65,
                80,
                75,
                100,
                100,
                70,
                55,
            ],
        )

        daily_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ])
        )

        story.append(daily_table)

    else:

        story.append(
            Paragraph(
                "No daily progress records found.",
                styles["Normal"],
            )
        )

    story.append(
        Spacer(1, 20)
    )

    # ========================================================
    # WEEKLY PROGRESS
    # ========================================================

    story.append(
        Paragraph(
            "Weekly Progress Reports",
            styles["Heading2"],
        )
    )

    if weekly_progress:

        weekly_data = [
            [
                "Week",
                "Work Completed",
                "Completion %",
                "Worker Hours",
                "Major Activities",
                "Delays",
                "Safety Incidents",
                "Status",
            ]
        ]

        for item in weekly_progress:

            weekly_data.append([
                f"{item.week_start or ''} to {item.week_end or ''}",
                item.work_completed or "",
                str(item.completion_percentage or 0),
                str(item.worker_hours or 0),
                item.major_activities or "",
                item.delays or "",
                item.safety_incidents or "",
                item.overall_status or "",
            ])

        weekly_table = Table(
            weekly_data,
            repeatRows=1,
            colWidths=[
                90,
                130,
                65,
                70,
                130,
                100,
                100,
                80,
            ],
        )

        weekly_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ])
        )

        story.append(weekly_table)

    else:

        story.append(
            Paragraph(
                "No weekly progress records found.",
                styles["Normal"],
            )
        )

    story.append(
        Spacer(1, 20)
    )

    # ========================================================
    # DELAYS
    # ========================================================

    story.append(
        Paragraph(
            "Delay Records",
            styles["Heading2"],
        )
    )

    if delays:

        delay_data = [
            [
                "Date",
                "Reason",
                "Duration",
                "Affected Work",
                "Impact",
            ]
        ]

        for item in delays:

            delay_data.append([
                str(item.delay_date or ""),
                item.reason or "",
                f"{item.duration_hours or 0} hrs",
                item.affected_work or "",
                item.impact or "",
            ])

        delay_table = Table(
            delay_data,
            repeatRows=1,
            colWidths=[
                80,
                150,
                70,
                150,
                220,
            ],
        )

        delay_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ])
        )

        story.append(delay_table)

    else:

        story.append(
            Paragraph(
                "No delay records found.",
                styles["Normal"],
            )
        )

    story.append(
        Spacer(1, 20)
    )

    # ========================================================
    # MILESTONES
    # ========================================================

    story.append(
        Paragraph(
            "Project Milestones",
            styles["Heading2"],
        )
    )

    if milestones:

        milestone_data = [
            [
                "Milestone",
                "Description",
                "Due Date",
                "Status",
            ]
        ]

        for item in milestones:

            milestone_data.append([
                item.title or "",
                item.description or "",
                str(item.due_date or ""),
                item.status or "",
            ])

        milestone_table = Table(
            milestone_data,
            repeatRows=1,
            colWidths=[
                180,
                300,
                100,
                100,
            ],
        )

        milestone_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ])
        )

        story.append(milestone_table)

    else:

        story.append(
            Paragraph(
                "No milestones found.",
                styles["Normal"],
            )
        )

    # --------------------------------------------------------
    # Build PDF
    # --------------------------------------------------------

    document.build(story)

    return FileResponse(
        path=file_path,
        filename=file_name,
        media_type="application/pdf",
    )


# ============================================================
# GENERATE PROJECT EXCEL REPORT
# ============================================================

@router.get("/project/{project_id}/excel")
def generate_project_excel(
    project_id: int,
    db: Session = Depends(get_db),
):
    # --------------------------------------------------------
    # Get project
    # --------------------------------------------------------

    project = (
        db.query(Project)
        .filter(Project.id == project_id)
        .first()
    )

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found",
        )

    # --------------------------------------------------------
    # Get Module 3 data
    # --------------------------------------------------------

    daily_progress = (
        db.query(DailyProgress)
        .filter(DailyProgress.project_id == project_id)
        .order_by(DailyProgress.report_date)
        .all()
    )

    weekly_progress = (
        db.query(WeeklyProgress)
        .filter(WeeklyProgress.project_id == project_id)
        .order_by(WeeklyProgress.week_start)
        .all()
    )

    delays = (
        db.query(DelayRecord)
        .filter(DelayRecord.project_id == project_id)
        .order_by(DelayRecord.delay_date)
        .all()
    )

    milestones = (
        db.query(ProjectMilestone)
        .filter(ProjectMilestone.project_id == project_id)
        .order_by(ProjectMilestone.due_date)
        .all()
    )

    # --------------------------------------------------------
    # Create output directory
    # --------------------------------------------------------

    output_directory = "generated_reports"

    os.makedirs(
        output_directory,
        exist_ok=True
    )

    file_name = (
        f"project_{project_id}_progress_report.xlsx"
    )

    file_path = os.path.join(
        output_directory,
        file_name
    )

    # --------------------------------------------------------
    # Create workbook
    # --------------------------------------------------------

    workbook = Workbook()

    # ========================================================
    # PROJECT SUMMARY SHEET
    # ========================================================

    summary_sheet = workbook.active
    summary_sheet.title = "Project Summary"

    summary_data = [
        ["Project Name", project.project_name],
        ["Project Code", project.project_code],
        ["Category", project.project_category],
        ["Location", project.location],
        ["Start Date", project.start_date],
        ["End Date", project.end_date],
        ["Budget", project.budget],
        ["Status", project.status],
    ]

    for row in summary_data:
        summary_sheet.append(row)

    # ========================================================
    # DAILY PROGRESS SHEET
    # ========================================================

    daily_sheet = workbook.create_sheet(
        "Daily Progress"
    )

    daily_headers = [
        "Date",
        "Work Category",
        "Activity",
        "Completion %",
        "Contractor",
        "Workers Present",
        "Workers Absent",
        "Machinery Used",
        "Materials Used",
        "Weather",
        "Safety Observation",
        "Quality Remarks",
        "Quality Verified",
        "Delay Hours",
        "Delay Reason",
        "Comments",
    ]

    daily_sheet.append(daily_headers)

    for item in daily_progress:

        daily_sheet.append([
            item.report_date,
            item.work_category,
            item.activity,
            item.completion_percentage,
            item.contractor_name,
            item.workers_present,
            item.workers_absent,
            item.machinery_used,
            item.materials_used,
            item.weather,
            item.safety_observation,
            item.quality_remarks,
            item.quality_verified,
            item.delay_hours,
            item.delay_reason,
            item.comments,
        ])

    # ========================================================
    # WEEKLY PROGRESS SHEET
    # ========================================================

    weekly_sheet = workbook.create_sheet(
        "Weekly Progress"
    )

    weekly_headers = [
        "Week Start",
        "Week End",
        "Work Completed",
        "Completion %",
        "Worker Hours",
        "Major Activities",
        "Delays",
        "Safety Incidents",
        "Overall Status",
    ]

    weekly_sheet.append(weekly_headers)

    for item in weekly_progress:

        weekly_sheet.append([
            item.week_start,
            item.week_end,
            item.work_completed,
            item.completion_percentage,
            item.worker_hours,
            item.major_activities,
            item.delays,
            item.safety_incidents,
            item.overall_status,
        ])

    # ========================================================
    # DELAY SHEET
    # ========================================================

    delay_sheet = workbook.create_sheet(
        "Delay Records"
    )

    delay_headers = [
        "Delay Date",
        "Reason",
        "Duration Hours",
        "Affected Work",
        "Impact",
    ]

    delay_sheet.append(delay_headers)

    for item in delays:

        delay_sheet.append([
            item.delay_date,
            item.reason,
            item.duration_hours,
            item.affected_work,
            item.impact,
        ])

    # ========================================================
    # MILESTONES SHEET
    # ========================================================

    milestone_sheet = workbook.create_sheet(
        "Milestones"
    )

    milestone_headers = [
        "Title",
        "Description",
        "Due Date",
        "Status",
    ]

    milestone_sheet.append(milestone_headers)

    for item in milestones:

        milestone_sheet.append([
            item.title,
            item.description,
            item.due_date,
            item.status,
        ])

    # ========================================================
    # FORMAT ALL SHEETS
    # ========================================================

    for sheet in workbook.worksheets:

        # Header formatting
        for cell in sheet[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # Cell formatting
        for row in sheet.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        # Auto-size columns
        for column_cells in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                try:
                    cell_length = len(
                        str(cell.value)
                    )

                    if cell_length > max_length:
                        max_length = cell_length

                except Exception:
                    pass

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                40
            )

        # Freeze header row
        sheet.freeze_panes = "A2"

    # --------------------------------------------------------
    # Save workbook
    # --------------------------------------------------------

    workbook.save(file_path)

    return FileResponse(
        path=file_path,
        filename=file_name,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
